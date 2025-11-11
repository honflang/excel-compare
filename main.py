import json
import os
import shutil
import sys
import traceback
from copy import copy
from dataclasses import dataclass
from datetime import datetime

import openpyxl
import pandas as pd
from openpyxl.comments import Comment
from openpyxl.styles import Font, PatternFill

def_config_file_name = 'config.json'
config = None
compared_sheet_name = []


@dataclass
class ColumnConfig:
    name: str = None
    sub: list[int] = None

    @classmethod
    def from_json(cls, json_data):
        if isinstance(json_data, str):
            return cls(name=json_data)
        else:
            return cls(name=json_data['name'],
                       sub=json_data.get('sub'))


@dataclass
class SheetConfig:
    unique_columns: list[ColumnConfig]
    compare_columns: list[ColumnConfig]
    index: int = None
    name: str = None
    skip_lines: int = 0  # 跳过的数据行数（除表头外也就是从第二行开始计算）

    @classmethod
    def from_json(cls, json_data):
        unique_columns = [ColumnConfig.from_json(unique_column) for unique_column in json_data['unique_columns']]
        compare_columns = [ColumnConfig.from_json(compare_column) for compare_column in json_data['compare_columns']]
        return cls(unique_columns=unique_columns,
                   compare_columns=compare_columns,
                   index=json_data.get('index'),
                   name=json_data.get('name'),
                   skip_lines=json_data.get('skip_lines') or 0
                   )


@dataclass
class Config:
    main_compare_file_path: str  # 主文件名
    sub_compare_file_path: str  # 副文件名
    sheets: list[SheetConfig]  # Sheet 配置
    output_path: str = ""  # 对比文件输出位置
    skip_both: bool = True
    result_file_ptah: str = None  # 对比结果文件名

    @classmethod
    def from_json(cls, json_data):
        sheets = [SheetConfig.from_json(sheet) for sheet in json_data['sheets']]
        return cls(
            main_compare_file_path=json_data['main_compare_file_path'],
            sub_compare_file_path=json_data['sub_compare_file_path'],
            output_path=json_data['output_path'],
            skip_both=json_data['skip_both'],
            sheets=sheets
        )


def compare():
    copy_file_shutil_copy()
    for sheet_config in config.sheets:
        compare_sheet(sheet_config)


def compare_sheet(sheet_config: SheetConfig):
    # TODO 是否可以避免重复加载保存结果文件
    sheet_index = sheet_config.index if (
            sheet_config.index is not None and sheet_config.index >= 0) else sheet_config.name
    if sheet_index is None:
        print(f"[Warn] Invalid sheet index : {sheet_index}")
        return None

    main_sheet = read_excel(config.main_compare_file_path, sheet_index)
    sub_sheet = read_excel(config.sub_compare_file_path, sheet_index)

    if main_sheet is None:
        print(f"[Warn] Sheet '{sheet_index}' not found in '{config.main_compare_file_path}'")
        return None

    if sub_sheet is None:
        print(f"[Warn] Sheet '{sheet_index}' not found in '{config.sub_compare_file_path}'")
        return None

    try:
        unique_columns_name = [col.name for col in sheet_config.unique_columns]
        missing_columns = set(unique_columns_name) - set(main_sheet.columns)
        if missing_columns:
            print(
                f"[Warn] File '{config.main_compare_file_path}' sheet[{sheet_index}] 缺少必要的列: {', '.join(missing_columns)}")
        missing_columns_2 = set(unique_columns_name) - set(sub_sheet.columns)
        if missing_columns_2:
            print(
                f"[Warn] File '{config.sub_compare_file_path}' sheet[{sheet_index}] 缺少必要的列: {', '.join(missing_columns_2)}")

        if missing_columns or missing_columns_2:
            return None

        main_sheet_data = get_sheet_data(main_sheet, sheet_config)
        sub_sheet_data = get_sheet_data(sub_sheet, sheet_config)
        workbook = openpyxl.load_workbook(config.result_file_ptah)

        if isinstance(sheet_index, str):
            result_sheet = workbook[sheet_index]
        else:
            result_sheet = workbook.worksheets[sheet_index]

        result_sheet_name = result_sheet.title
        if result_sheet_name in compared_sheet_name:
            print(f"[Warn] Sheet '{result_sheet_name}' already compared")
            return None

        compared_sheet_name.append(result_sheet_name)

        result_sheet.insert_cols(1, 2)
        result_sheet['A1'].value = list_to_str(unique_columns_name)
        result_sheet['B1'].value = "对比结果"
        result_sheet.column_dimensions['A'].width = 40
        result_sheet.column_dimensions['B'].width = 20

        copy_cell_style(result_sheet['C1'], result_sheet['A1'])
        copy_cell_style(result_sheet['C1'], result_sheet['B1'])

        match_to_multiple_lines = []
        df_count = 0
        eq_count = 0
        repetition_count = 0
        m_overflowing_count = 0
        m_max_len = len(main_sheet_data["unique_values"])
        for i in range(0, m_max_len):
            if i < sheet_config.skip_lines:
                continue
            mu = main_sheet_data["unique_values"][i]
            result_sheet.cell(row=i + 2, column=1, value=list_to_str(mu))

            if main_sheet_data["unique_values"].count(mu) > 1:
                result_sheet.cell(row=i + 2, column=2, value="当前表重复")
                repetition_count += 1
                continue

            try:
                index = sub_sheet_data["unique_values"].index(mu)
                if sub_sheet_data["unique_values"].count(mu) > 1:
                    match_to_multiple_lines.append(i)
                    result_sheet.cell(row=i + 2, column=2, value="他表匹配到多列")
                else:
                    if main_sheet_data["compare_values"][i] == sub_sheet_data["compare_values"][index]:
                        eq_count += 1
                        result_sheet.cell(row=i + 2, column=2, value="一致")
                    else:
                        result_sheet.cell(row=i + 2, column=2, value="不一致")
                        df_count += 1
                        for j in range(0, len(main_sheet_data["compare_values"][i])):
                            main_sheet_data_compare_value = main_sheet_data["compare_values"][i][j]
                            sub_sheet_data_compare_value = sub_sheet_data["compare_values"][index][j]
                            if main_sheet_data_compare_value != sub_sheet_data_compare_value:
                                cell = result_sheet.cell(row=i + 2, column=main_sheet_data["compare_indices"][j] + 3)
                                cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
                                cell.comment = Comment(
                                    f"当前值：{main_sheet_data_compare_value or '空'}\r\n他表值：{sub_sheet_data_compare_value or '空'}",
                                    "<System>")
            except ValueError:
                result_sheet.cell(row=i + 2, column=2, value="他表缺失")
                m_overflowing_count += 1

        s_overflowing_count = 0
        s_max_len = len(sub_sheet_data["unique_values"])
        for i in range(0, s_max_len):
            if config.skip_both and i < sheet_config.skip_lines:
                continue
            su = sub_sheet_data["unique_values"][i]
            try:
                main_sheet_data["unique_values"].index(su)
            except ValueError:
                result_sheet.cell(row=s_overflowing_count + m_max_len + 2, column=1, value=list_to_str(su))
                result_sheet.cell(row=s_overflowing_count + m_max_len + 2, column=2, value="当前表缺失")
                s_overflowing_count += 1

        # print(f"match_to_multiple_lines : {match_to_multiple_lines}")
        # print(f"m_overflowing_lines : {m_overflowing_lines}")

        total_line_index = m_max_len + s_overflowing_count
        result_sheet['A' + str(total_line_index + 3)].value = "总行数：" + str(
            total_line_index - sheet_config.skip_lines)
        result_sheet['A' + str(total_line_index + 4)].value = "一致数：" + str(eq_count)
        result_sheet['A' + str(total_line_index + 5)].value = "不一致数：" + str(df_count)
        result_sheet['A' + str(total_line_index + 6)].value = "他表缺失数：" + str(m_overflowing_count)
        result_sheet['A' + str(total_line_index + 7)].value = "当前表缺失数：" + str(s_overflowing_count)
        result_sheet['A' + str(total_line_index + 8)].value = "当前表重复行数：" + str(repetition_count)
        result_sheet['A' + str(total_line_index + 9)].value = "他表匹配到多列数：" + str(len(match_to_multiple_lines))

        # result_sheet['A' + str(total_line_index + 3)].font = Font(color="FF0000")
        result_sheet['A' + str(total_line_index + 4)].font = Font(color="00FF00")
        result_sheet['A' + str(total_line_index + 5)].font = Font(color="FF0000")
        result_sheet['A' + str(total_line_index + 6)].font = Font(color="FF0000")
        result_sheet['A' + str(total_line_index + 7)].font = Font(color="FF0000")
        result_sheet['A' + str(total_line_index + 8)].font = Font(color="FF0000")
        result_sheet['A' + str(total_line_index + 9)].font = Font(color="FF0000")

        workbook.save(config.result_file_ptah)

        print(f"[Info] Successfully compared sheet '{result_sheet_name}'")
    except KeyError as e:
        print(f"A required column was not found. Please check your column names. Details: {e}")
    except Exception as e:
        traceback.print_exc()
        print(f"An unexpected error occurred: {e}")


def list_to_str(l):
    if l is None:
        return ""
    else:
        return '/'.join([str(x) if x is not None else '空' for x in l])


def get_sheet_data(sheet, sheet_config: SheetConfig):
    if sheet is None or sheet_config is None:
        return None
    sheet_data = {"unique_values": [], "compare_values": []}
    compare_indices = []
    column_data = sheet[sheet_config.unique_columns[0].name]
    for i in range(len(column_data)):
        unique_value = []
        compare_value = []
        for j in range(len(sheet_config.unique_columns)):
            # unique_value.append(nan_as_none(sheet[sheet_config.unique_columns[j]][i]))
            unique_value.append(get_cell_value(sheet, i, sheet_config.unique_columns[j]))
        for k in range(len(sheet_config.compare_columns)):
            if i == 0:
                compare_indices.append(sheet.columns.get_loc(sheet_config.compare_columns[k].name))
            # compare_value.append(nan_as_none(sheet[sheet_config.compare_columns[k]][i]))
            compare_value.append(get_cell_value(sheet, i, sheet_config.compare_columns[k]))
        sheet_data["unique_values"].append(unique_value)
        sheet_data["compare_values"].append(compare_value)

    sheet_data["compare_indices"] = compare_indices
    return sheet_data


def get_cell_value(sheet, line: int, column_config: ColumnConfig):
    val = nan_as_none(sheet[column_config.name][line])
    if val is None:
        return None
    else:
        if column_config.sub is not None and len(column_config.sub) > 0:
            if len(column_config.sub) == 1:
                return str(val)[:column_config.sub[0]]
            if len(column_config.sub) == 2:
                if column_config.sub[0] is None:
                    return str(val)[column_config.sub[1]:]
                else:
                    return str(val)[column_config.sub[0]:column_config.sub[1]]
        return val


def nan_as_none(value):
    return None if pd.isna(value) else value


def rgb_to_hex(r, g, b):
    return f'{r:02X}{g:02X}{b:02X}'


def complete_file_path(path=""):
    if path is None:
        return "./"
    if not path.endswith('/') or path.endswith('\\'):
        path += '/'
    return path


def read_excel(filename, index):
    try:
        return pd.read_excel(filename, sheet_name=index)
    except Exception as e:
        # print(f"  --  An unexpected error occurred: {e}")
        return None


def get_filename_from_path(full_path):
    """
    使用 os.path.basename() 从完整路径中获取文件名。

    Args:
        full_path (str): 完整的文件路径字符串。

    Returns:
        str: 文件名部分。
    """
    if not isinstance(full_path, str):
        raise TypeError("输入必须是字符串。")

    return os.path.basename(full_path)


def load_config(file_path=None):
    if file_path is None:
        file_path = def_config_file_name
    print(f"Loading config from file : {file_path}")
    with open(file_path, 'r', encoding='utf-8') as f:
        json_data = json.load(f)
    new_config = Config.from_json(json_data)
    main_compare_file_name = get_filename_from_path(new_config.main_compare_file_path)
    new_config.result_file_ptah = (complete_file_path(new_config.output_path) + "compared-"
                                   + main_compare_file_name.split(
                '.')[0]
                                   + datetime.now().strftime(
                "-%Y%m%d%H%M%S.")
                                   + main_compare_file_name.split(
                '.')[1])
    # print(f"Loaded config : {new_config}")
    return new_config


def copy_file_shutil_copy():
    try:
        shutil.copy(config.main_compare_file_path, config.result_file_ptah)
    except FileNotFoundError:
        print(f"  --  Error: 源文件 '{config.main_compare_file_path}' 未找到。")
    except IsADirectoryError:
        print(f"  --  Error: 目标 '{config.result_file_ptah}' 是一个目录，但你指定了文件作为目标。")
    except Exception as e:
        print(f"  --  复制文件时发生错误: {e}")


def copy_cell_style(source_cell, target_cell):
    """
    完全复制源单元格的样式到目标单元格
    """
    if source_cell.has_style:
        target_cell.font = Font(
            name=source_cell.font.name,
            size=source_cell.font.size,
            bold=source_cell.font.bold,
            italic=source_cell.font.italic,
            vertAlign=source_cell.font.vertAlign,
            underline=source_cell.font.underline,
            strike=source_cell.font.strike,
            color=source_cell.font.color if source_cell.font.color else None
        ) if source_cell.font else None

        target_cell.border = copy(source_cell.border)
        target_cell.fill = copy(source_cell.fill)
        target_cell.number_format = source_cell.number_format
        target_cell.protection = copy(source_cell.protection)
        target_cell.alignment = copy(source_cell.alignment)


if __name__ == "__main__":

    # sys.argv[0] 是脚本本身的名称
    try:
        if len(sys.argv) > 1:
            config = load_config(sys.argv[1])
        else:
            config = load_config()
    except Exception as ex:
        print(f"  --  Load config file failed")
        print(ex)
        traceback.print_exc()
        sys.exit(1)
    try:
        compare()
    except Exception as ex:
        print(ex)
        traceback.print_exc()
        sys.exit(1)
